"""
rag_engine.py
custom_qa.json을 마스터 데이터베이스로 사용하는 RAG 엔진.
기본 엑셀(dormitory_guide_v2.xlsx)은 최초 구동 시 custom_qa.json으로 자동 이관됩니다.
"""

import os
import json
from openpyxl import load_workbook

EXCEL_FILENAME = "dormitory_guide_v2.xlsx"
SHEET_NAME = "기숙사_운영_데이터"
CUSTOM_QA_FILENAME = "custom_qa.json"

# 컬럼 순서 (A=카테고리, B=질문, C=답변)
COL_CATEGORY = 0
COL_QUESTION = 1
COL_ANSWER   = 2

# 캐시
_cache = {"data": None, "mtime": None}


def _get_excel_path():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, EXCEL_FILENAME)


def _get_custom_qa_path():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, CUSTOM_QA_FILENAME)


def save_data(data) -> bool:
    """마스터 JSON 데이터베이스(custom_qa.json)에 데이터를 씁니다."""
    path = _get_custom_qa_path()
    try:
        clean_data = []
        for item in data:
            clean_data.append({
                "category": str(item.get("category", "기타")).strip(),
                "question": str(item.get("question", "")).strip(),
                "answer": str(item.get("answer", "")).strip(),
                "is_custom": True  # 모든 데이터를 이제 웹에서 수정 가능하게 함
            })
        with open(path, "w", encoding="utf-8") as f:
            json.dump(clean_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"[RAG WARNING] custom_qa.json 저장 실패: {e}")
        return False


def load_data() -> list[dict]:
    """마스터 JSON 데이터를 읽어옵니다. 없을 경우 기본 엑셀에서 이관합니다."""
    global _cache
    custom_path = _get_custom_qa_path()
    excel_path = _get_excel_path()
    
    custom_mtime = os.path.getmtime(custom_path) if os.path.exists(custom_path) else 0

    if _cache["data"] is None or _cache["mtime"] != custom_mtime:
        # 1) custom_qa.json이 존재하는 경우
        if os.path.exists(custom_path):
            try:
                with open(custom_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    # 하위 호환성을 위해 속성 보정
                    for item in data:
                        item["is_custom"] = True
                    _cache["data"] = data
                    _cache["mtime"] = custom_mtime
                    print(f"[RAG] JSON 마스터 데이터 로드 완료: {len(data)}개")
                    return _cache["data"]
            except Exception as e:
                print(f"[RAG] custom_qa.json 로드 실패, 엑셀 이관을 시도합니다: {e}")

        # 2) custom_qa.json이 없거나 깨진 경우: 엑셀 파일에서 마이그레이션 진행
        excel_data = []
        if os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path, read_only=True, data_only=True)
                # 시트명이 다르면 첫 번째 시트 로드
                sheet_name = SHEET_NAME if SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

                for row in rows[1:]:  # 첫 행은 헤더
                    if len(row) < 3:
                        continue
                    cat = str(row[COL_CATEGORY]).strip() if row[COL_CATEGORY] else "기타"
                    q   = str(row[COL_QUESTION]).strip() if row[COL_QUESTION] else ""
                    a   = str(row[COL_ANSWER]).strip()   if row[COL_ANSWER]   else ""
                    if q and a and cat != "None":
                        excel_data.append({
                            "category": cat,
                            "question": q,
                            "answer": a,
                            "is_custom": True
                        })
                # JSON으로 마스터화 저장
                save_data(excel_data)
                _cache["data"] = excel_data
                _cache["mtime"] = os.path.getmtime(custom_path) if os.path.exists(custom_path) else 0
                print(f"[RAG] 엑셀에서 JSON 마스터 데이터 이관 완료: {len(excel_data)}개")
            except Exception as e:
                print(f"[RAG] 엑셀 이관 실패: {e}")
                _cache["data"] = []
        else:
            _cache["data"] = []

    return _cache["data"]


def get_categories() -> list:
    data = load_data()
    seen = []
    for item in data:
        if item["category"] not in seen:
            seen.append(item["category"])
    return seen


def search(query: str, top_k: int = 3) -> list[dict]:
    """질문과 가장 관련 있는 Q&A top_k개 반환."""
    data = load_data()
    if not data:
        return []

    query_lower = query.lower()

    scores = []
    for item in data:
        q_text   = item["question"].lower()
        a_text   = item["answer"].lower()
        combined = q_text + " " + a_text

        # 2글자 이상 슬라이딩 부분 문자열 매칭
        char_score = 0
        for length in range(2, min(len(query_lower) + 1, 8)):
            for i in range(len(query_lower) - length + 1):
                sub = query_lower[i:i + length]
                if sub in combined:
                    char_score += length

        scores.append((char_score, item))

    # 점수 내림차순 정렬 후 상위 k개
    scores.sort(key=lambda x: x[0], reverse=True)
    results = [item for score, item in scores[:top_k] if score > 0]
    return results


def get_all_qa() -> list[dict]:
    return load_data()


def add_qa(category, question, answer) -> bool:
    """Q&A 질문을 추가합니다."""
    data = load_data()
    
    # 중복 질문인 경우 덮어쓰기
    updated = False
    for item in data:
        if item["question"].strip() == question.strip():
            item["category"] = category.strip()
            item["answer"] = answer.strip()
            updated = True
            break
            
    if not updated:
        data.append({
            "category": category.strip(),
            "question": question.strip(),
            "answer": answer.strip(),
            "is_custom": True
        })
        
    success = save_data(data)
    reload()
    return success


def update_qa(old_question, category, new_question, answer) -> bool:
    """질문 내용을 수정합니다."""
    data = load_data()
    
    updated = False
    for item in data:
        if item["question"].strip() == old_question.strip():
            item["category"] = category.strip()
            item["question"] = new_question.strip()
            item["answer"] = answer.strip()
            updated = True
            break
            
    if not updated:
        return False
        
    success = save_data(data)
    reload()
    return success


def delete_qa(question_text) -> bool:
    """Q&A 질문을 삭제합니다."""
    data = load_data()
    original_len = len(data)
    
    data = [item for item in data if item["question"].strip() != question_text.strip()]
    
    if len(data) == original_len:
        return False
        
    success = save_data(data)
    reload()
    return success


def reload():
    global _cache
    _cache = {"data": None, "mtime": None}
    load_data()
    print("[RAG] Cache refreshed")
