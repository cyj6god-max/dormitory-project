"""
app.py — 기숙사 챗봇 Flask 웹 서버 (카테고리 필터 제거, 채팅 전용)
"""

import sys
import os
import shutil

# Windows 콘솔 인코딩 강제 UTF-8 설정
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass
os.environ.setdefault('PYTHONUTF8', '1')

from flask import Flask, request, jsonify, render_template, session, redirect, url_for, send_file
from dotenv import load_dotenv
from openai import OpenAI
import rag_engine
import json
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# app.py 위치 기준 절대 경로로 .env 로드
base_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(base_dir, ".env"))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024
app.config["JSON_AS_ASCII"] = False  # 한국어 JSON 응답 깨짐 방지
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dormitory_secret_key_12345")

api_key = os.getenv("OPENAI_API_KEY")
client = None
if api_key:
    try:
        client = OpenAI(api_key=api_key)
    except Exception as e:
        print(f"OpenAI Client 생성 실패: {e}")

EXCEL_FILENAME = "dormitory_guide_v2.xlsx"
CONFIG_FILE = "admin_config.json"
LOGS_FILE = "chat_logs.json"

def _get_logs_path():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, LOGS_FILE)

def load_logs() -> list:
    path = _get_logs_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"로그 로드 오류: {e}")
        return []

def save_logs(logs):
    path = _get_logs_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"로그 저장 실패 (Vercel 환경일 수 있음): {e}")

def write_chat_log(query, answer, found):
    logs = load_logs()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    logs.append({
        "timestamp": now_str,
        "query": query.strip(),
        "answer": answer.strip(),
        "found": found
    })
    
    # 3개월(90일) 보관 기간 필터링
    limit_date = datetime.now() - timedelta(days=90)
    filtered_logs = []
    for log in logs:
        try:
            log_time = datetime.strptime(log["timestamp"], "%Y-%m-%d %H:%M:%S")
            if log_time >= limit_date:
                filtered_logs.append(log)
        except Exception:
            filtered_logs.append(log)
            
    save_logs(filtered_logs)

def get_admin_password():
    if not os.path.exists(CONFIG_FILE):
        return "8002"
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data.get("password", "8002")
    except:
        return "8002"

def save_admin_password(new_password):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({"password": new_password}, f)


# ── 페이지 ──────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/admin")
def admin():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
    qa_list = rag_engine.get_all_qa()
    return render_template("admin.html", qa_list=qa_list, total=len(qa_list))


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if session.get("admin_logged_in"):
        return redirect(url_for("admin"))
    
    error = None
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        correct_password = get_admin_password()
        if password == correct_password:
            session["admin_logged_in"] = True
            return redirect(url_for("admin"))
        else:
            error = "비밀번호가 올바르지 않습니다."
            
    return render_template("login.html", error=error)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))


@app.route("/admin/change-password", methods=["POST"])
def change_password():
    if not session.get("admin_logged_in"):
        return jsonify({"error": "로그인이 필요합니다."}), 401
        
    data = request.get_json()
    if not data:
        return jsonify({"error": "올바르지 않은 요청 데이터입니다."}), 400
        
    current_pwd = data.get("current_password", "").strip()
    new_pwd = data.get("new_password", "").strip()
    
    if not current_pwd or not new_pwd:
        return jsonify({"error": "현재 비밀번호와 새 비밀번호를 모두 입력해 주세요."}), 400
        
    if current_pwd != get_admin_password():
        return jsonify({"error": "현재 비밀번호가 일치하지 않습니다."}), 400
        
    save_admin_password(new_pwd)
    session.pop("admin_logged_in", None)
    return jsonify({"success": True, "message": "비밀번호가 성공적으로 변경되었습니다. 다시 로그인해 주세요."})


@app.route("/admin/find-password", methods=["POST"])
def find_password():
    data = request.get_json()
    if not data:
        return jsonify({"error": "올바르지 않은 요청 데이터입니다."}), 400
        
    name = data.get("name", "").strip()
    init_pwd = data.get("init_password", "").strip()
    
    if not name or not init_pwd:
        return jsonify({"error": "이름과 초기 비밀번호를 모두 입력해 주세요."}), 400
        
    if name == "채유진" and init_pwd == "8002":
        current_password = get_admin_password()
        return jsonify({"success": True, "password": current_password})
    else:
        return jsonify({"error": "관리자 정보가 일치하지 않습니다."}), 400


# ── API ──────────────────────────────────────────

@app.route("/api/chat", methods=["POST"])
def chat():
    data = request.get_json()
    if not data or not data.get("message", "").strip():
        return jsonify({"error": "질문을 입력해 주세요."}), 400

    user_message = data["message"].strip()

    # 1) 엑셀에서 관련 Q&A 검색
    results = rag_engine.search(user_message, top_k=3)

    # 2) 관련 내용이 없으면 자료 없음 안내
    if not results:
        no_ref_answer = "적절한 답변을 찾지 못했습니다. 관리자에게 문의해 주세요. (소통폰 : 010-2629-8002)"
        write_chat_log(user_message, no_ref_answer, False)
        return jsonify({
            "answer": no_ref_answer,
            "references": [],
            "found": False,
        })

    # 3) 찾은 내용을 컨텍스트로 GPT에 전달
    context_text = ""
    for i, r in enumerate(results, 1):
        context_text += f"\n[자료 {i}] 카테고리: {r['category']}\n질문: {r['question']}\n답변: {r['answer']}\n"

    system_prompt = """당신은 기숙사 거주자의 질문에 답변하는 도우미입니다.
반드시 아래 규칙을 따르세요:
1. 아래 '기숙사 안내 자료'의 내용만을 참고해서 답변하세요.
2. 자료에 없는 내용은 절대 추측하거나 지어내지 마세요.
3. 자료에 답변이 없으면 "적절한 답변을 찾지 못했습니다. 관리자에게 문의해 주세요. (소통폰 : 010-2629-8002)"라고만 말하세요.
4. 답변은 친절하고 명확하게 한국어로 작성하세요.
5. 중요한 수치(날짜, 시간, 벌점 등)는 정확히 그대로 인용하세요."""

    user_prompt = f"""기숙사 안내 자료:
{context_text}

거주자 질문: {user_message}

위 자료를 바탕으로 질문에 답변해 주세요."""

    try:
        if not client:
            return jsonify({"error": "OpenAI API 키가 설정되지 않았거나 올바르지 않습니다. 버셀 설정에서 환경 변수를 등록해 주세요."}), 500
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.2,
            max_tokens=500,
        )
        answer = response.choices[0].message.content.strip()
    except Exception as e:
        return jsonify({"error": f"AI 서버 오류: {str(e)}"}), 500

    references = [{"category": r["category"], "question": r["question"]} for r in results]
    write_chat_log(user_message, answer, True)
    return jsonify({"answer": answer, "references": references, "found": True})


@app.route("/api/qa/upload", methods=["POST"])
def upload_qa_excel():
    if not session.get("admin_logged_in"):
        return jsonify({"error": "로그인이 필요합니다."}), 401
        
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "엑셀 파일(.xlsx, .xls)만 업로드 가능합니다."}), 400

    base_dir = os.path.dirname(os.path.abspath(__file__))
    save_path = os.path.join(base_dir, EXCEL_FILENAME)
    temp_path = save_path + ".tmp.xlsx"
    backup_path = save_path + ".bak"

    # 1) 임시 파일로 저장
    try:
        file.save(temp_path)
    except Exception as e:
        return jsonify({"error": f"파일 저장 오류: {str(e)}"}), 500

    # 2) 시트 구조 검증
    try:
        from openpyxl import load_workbook
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        if "기숙사_운영_데이터" not in wb.sheetnames:
            wb.close()
            os.remove(temp_path)
            return jsonify({"error": "시트 이름이 '기숙사_운영_데이터'인 시트가 존재해야 합니다. (템플릿을 참고하세요)"}), 400
        wb.close()
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": f"엑셀 읽기 오류: {str(e)}"}), 400

    # 3) 기존 파일 백업 후 교체
    try:
        if os.path.exists(save_path):
            import shutil
            shutil.copy2(save_path, backup_path)
        os.replace(temp_path, save_path)
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": f"서버 파일 저장 권한 오류: {str(e)}"}), 500

    # 4) 캐시 갱신
    try:
        rag_engine.reload()
        count = len(rag_engine.get_all_qa())
        return jsonify({"success": True, "message": f"업로드 완료! 총 {count}개의 Q&A가 반영되었습니다.", "count": count})
    except Exception as e:
        return jsonify({"error": f"파일 처리 후 캐시 갱신 실패: {str(e)}"}), 500

@app.route("/api/qa/download", methods=["GET"])
def download_qa_excel():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
        
    qa_list = rag_engine.get_all_qa()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "기숙사_운영_데이터"
    
    ws.append(["카테고리", "질문", "답변"])
    
    for item in qa_list:
        ws.append([
            item.get("category", "기타"),
            item.get("question", ""),
            item.get("answer", "")
        ])
        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val = str(cell.value or '')
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="dormitory_guide_v2.xlsx"
    )

@app.route("/api/qa/add", methods=["POST"])
def add_qa():
    if not session.get("admin_logged_in"):
        return jsonify({"error": "로그인이 필요합니다."}), 401
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "올바르지 않은 요청 데이터입니다."}), 400
        
    category = data.get("category", "").strip()
    question = data.get("question", "").strip()
    answer = data.get("answer", "").strip()
    
    if not category or not question or not answer:
        return jsonify({"error": "카테고리, 질문, 답변을 모두 입력해 주세요."}), 400
        
    success = rag_engine.add_custom_qa(category, question, answer)
    if success:
        return jsonify({"success": True, "message": "Q&A가 성공적으로 추가되었습니다."})
    else:
        return jsonify({"success": True, "message": "Q&A가 추가되었습니다. (임시 환경 저장)"})

@app.route("/api/qa/delete", methods=["POST"])
def delete_qa():
    if not session.get("admin_logged_in"):
        return jsonify({"error": "로그인이 필요합니다."}), 401
        
    data = request.get_json()
    if not data or not data.get("question", "").strip():
        return jsonify({"error": "삭제할 질문을 지정해 주세요."}), 400
        
    question = data["question"].strip()
    success = rag_engine.delete_custom_qa(question)
    
    if success:
        return jsonify({"success": True, "message": "Q&A가 성공적으로 삭제되었습니다."})
    else:
        return jsonify({"error": "삭제할 Q&A를 찾지 못했거나 기본 제공 데이터(엑셀)는 삭제할 수 없습니다."}), 400

@app.route("/api/logs/daily", methods=["GET"])
def get_daily_logs():
    if not session.get("admin_logged_in"):
        return jsonify({"error": "로그인이 필요합니다."}), 401
        
    logs = load_logs()
    
    daily_counts = {}
    for log in logs:
        try:
            date_str = log["timestamp"].split(" ")[0]
            daily_counts[date_str] = daily_counts.get(date_str, 0) + 1
        except Exception:
            continue
            
    sorted_daily = []
    for d in sorted(daily_counts.keys(), reverse=True):
        sorted_daily.append({"date": d, "count": daily_counts[d]})
        
    return jsonify({"success": True, "data": sorted_daily})

@app.route("/api/logs/download", methods=["GET"])
def download_logs_excel():
    if not session.get("admin_logged_in"):
        return redirect(url_for("admin_login"))
        
    logs = load_logs()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "질문 내역 로그"
    
    headers = ["질문 일시", "사용자 질문", "챗봇 답변", "자료 매칭 여부"]
    ws.append(headers)
    
    for log in reversed(logs):
        ws.append([
            log.get("timestamp", ""),
            log.get("query", ""),
            log.get("answer", ""),
            "예" if log.get("found", False) else "아니오"
        ])
        
    for col in ws.columns:
        max_len = 0
        for cell in col:
            # None 방지 및 한글/영문 길이 보정
            val = str(cell.value or '')
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="chat_history_logs.xlsx"
    )



if __name__ == "__main__":
    print("=" * 50)
    print("  기숙사 챗봇 서버 시작")
    print("  거주자 화면: http://localhost:5000")
    print("  관리자 화면: http://localhost:5000/admin")
    print("=" * 50)
    app.run(debug=True, host="0.0.0.0", port=5000)
